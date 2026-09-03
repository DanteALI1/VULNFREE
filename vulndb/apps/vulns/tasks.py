"""
Celery sync tasks for NVD, CISA KEV and БДУ ФСТЭК.

Порядок NVD→KEV (критично):
  sync_nvd() всегда в конце своего выполнения (успех / demo-fallback / exception)
  — если kev_enabled=True — синхронно вызывает sync_kev().

  tick_sync_schedules: если nvd_enabled — НЕ планирует отдельный KEV-прогон
  (KEV подтягивается внутри sync_nvd). KEV отдельно только когда nvd_enabled=False.

  wizard finish: если nvd_enabled → только sync_nvd.delay();
  иначе если kev_enabled → sync_kev.delay().
"""

from __future__ import annotations

import logging
import time
from datetime import timedelta
from io import BytesIO
from typing import Any

from celery import shared_task
from django.utils import timezone
from django.utils.dateparse import parse_datetime

from vulndb.apps.vulns.http_utils import request_with_retries
from vulndb.apps.vulns.models import SyncState, Vulnerability

logger = logging.getLogger(__name__)

NVD_URL = "https://services.nvd.nist.gov/rest/json/cves/2.0"
KEV_URL = "https://www.cisa.gov/sites/default/files/feeds/known_exploited_vulnerabilities.json"
NVD_PAGE_SIZE = 100
NVD_MAX_RECORDS = 500


def _settings():
    from vulndb.apps.core.models import SystemSettings

    return SystemSettings.load()


def _severity_from_score(score: float | None) -> str:
    if score is None:
        return Vulnerability.Severity.UNKNOWN
    if score >= 9.0:
        return Vulnerability.Severity.CRITICAL
    if score >= 7.0:
        return Vulnerability.Severity.HIGH
    if score >= 4.0:
        return Vulnerability.Severity.MEDIUM
    if score > 0:
        return Vulnerability.Severity.LOW
    return Vulnerability.Severity.NONE


def _extract_cvss(metrics: dict) -> tuple[float | None, dict, dict, dict, dict]:
    v31 = v30 = v2 = v40 = {}
    score = None
    for key, bucket in (
        ("cvssMetricV31", "v31"),
        ("cvssMetricV30", "v30"),
        ("cvssMetricV2", "v2"),
        ("cvssMetricV40", "v40"),
    ):
        items = metrics.get(key) or []
        if not items:
            continue
        primary = next((i for i in items if i.get("type") == "Primary"), items[0])
        data = primary.get("cvssData") or {}
        if bucket == "v31":
            v31 = data
            score = score if score is not None else data.get("baseScore")
        elif bucket == "v30":
            v30 = data
            score = score if score is not None else data.get("baseScore")
        elif bucket == "v2":
            v2 = data
            score = score if score is not None else data.get("baseScore")
        else:
            v40 = data
            score = score if score is not None else data.get("baseScore")
    return score, v31, v30, v2, v40


def _upsert_cve_from_nvd(item: dict) -> Vulnerability:
    cve = item.get("cve") or item
    vuln_id = cve.get("id") or ""
    descs = cve.get("descriptions") or []
    title = ""
    description = ""
    for d in descs:
        if d.get("lang") == "en":
            description = d.get("value") or ""
            title = description[:240]
            break
    if not title and descs:
        description = descs[0].get("value") or ""
        title = description[:240]

    score, v31, v30, v2, v40 = _extract_cvss(cve.get("metrics") or {})
    weaknesses = []
    for w in cve.get("weaknesses") or []:
        for desc in w.get("description") or []:
            if desc.get("value"):
                weaknesses.append(desc["value"])
    cpes: list[str] = []
    for cfg in cve.get("configurations") or []:
        for node in cfg.get("nodes") or []:
            for m in node.get("cpeMatch") or []:
                if m.get("criteria"):
                    cpes.append(m["criteria"])
    refs = [r.get("url") for r in (cve.get("references") or []) if r.get("url")]

    published = parse_datetime(cve.get("published") or "") if cve.get("published") else None
    modified = parse_datetime(cve.get("lastModified") or "") if cve.get("lastModified") else None
    if published and timezone.is_naive(published):
        published = timezone.make_aware(published, timezone.get_current_timezone())
    if modified and timezone.is_naive(modified):
        modified = timezone.make_aware(modified, timezone.get_current_timezone())

    obj, _ = Vulnerability.objects.update_or_create(
        vuln_id=vuln_id,
        defaults={
            "record_type": Vulnerability.RecordType.CVE,
            "title": title or vuln_id,
            "description_nvd": description,
            "severity": _severity_from_score(float(score) if score is not None else None),
            "cvss_score": float(score) if score is not None else None,
            "cvss_v31": v31,
            "cvss_v30": v30,
            "cvss_v2": v2,
            "cvss_v40": v40,
            "cwe": weaknesses,
            "cpe": cpes[:200],
            "references": refs[:200],
            "published_at": published,
            "modified_at": modified or timezone.now(),
            "raw_nvd": item,
        },
    )
    return obj


DEMO_CVES = [
    {
        "cve": {
            "id": "CVE-2024-0001",
            "published": "2024-01-15T00:00:00.000",
            "lastModified": "2024-02-01T00:00:00.000",
            "descriptions": [
                {
                    "lang": "en",
                    "value": ("Demo critical remote code execution in ExampleServer auth module."),
                }
            ],
            "metrics": {
                "cvssMetricV31": [
                    {
                        "type": "Primary",
                        "cvssData": {
                            "version": "3.1",
                            "baseScore": 9.8,
                            "baseSeverity": "CRITICAL",
                            "vectorString": "CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:H/I:H/A:H",
                        },
                    }
                ]
            },
            "weaknesses": [{"description": [{"lang": "en", "value": "CWE-287"}]}],
            "references": [{"url": "https://nvd.nist.gov/vuln/detail/CVE-2024-0001"}],
        }
    },
    {
        "cve": {
            "id": "CVE-2024-0002",
            "published": "2024-03-10T00:00:00.000",
            "lastModified": "2024-03-20T00:00:00.000",
            "descriptions": [
                {
                    "lang": "en",
                    "value": "Demo high-severity SQL injection in ExampleCMS admin panel search.",
                }
            ],
            "metrics": {
                "cvssMetricV31": [
                    {
                        "type": "Primary",
                        "cvssData": {
                            "version": "3.1",
                            "baseScore": 8.1,
                            "baseSeverity": "HIGH",
                            "vectorString": "CVSS:3.1/AV:N/AC:L/PR:L/UI:N/S:U/C:H/I:H/A:N",
                        },
                    }
                ]
            },
            "weaknesses": [{"description": [{"lang": "en", "value": "CWE-89"}]}],
            "references": [{"url": "https://nvd.nist.gov/vuln/detail/CVE-2024-0002"}],
        }
    },
    {
        "cve": {
            "id": "CVE-2023-9999",
            "published": "2023-11-01T00:00:00.000",
            "lastModified": "2023-12-01T00:00:00.000",
            "descriptions": [
                {
                    "lang": "en",
                    "value": "Demo medium path traversal in ExampleFileShare download endpoint.",
                }
            ],
            "metrics": {
                "cvssMetricV31": [
                    {
                        "type": "Primary",
                        "cvssData": {
                            "version": "3.1",
                            "baseScore": 5.3,
                            "baseSeverity": "MEDIUM",
                            "vectorString": "CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:L/I:N/A:N",
                        },
                    }
                ]
            },
            "weaknesses": [{"description": [{"lang": "en", "value": "CWE-22"}]}],
            "references": [{"url": "https://nvd.nist.gov/vuln/detail/CVE-2023-9999"}],
        }
    },
]


def seed_demo_cves() -> int:
    if Vulnerability.objects.exists():
        return 0
    count = 0
    for item in DEMO_CVES:
        _upsert_cve_from_nvd(item)
        count += 1
    return count


@shared_task(name="vulndb.apps.vulns.tasks.sync_kev")
def sync_kev() -> dict[str, Any]:
    state = SyncState.get_or_create_source(SyncState.Source.KEV)
    state.mark_running()
    try:
        resp = request_with_retries("GET", KEV_URL, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        vulns = data.get("vulnerabilities") or []
        state.mark_progress(0, len(vulns))
        synced = 0
        for entry in vulns:
            cve_id = entry.get("cveID") or entry.get("cveId") or ""
            if not cve_id:
                continue
            obj, created = Vulnerability.objects.get_or_create(
                vuln_id=cve_id,
                defaults={
                    "record_type": Vulnerability.RecordType.CVE,
                    "title": entry.get("vulnerabilityName") or cve_id,
                    "description_nvd": entry.get("shortDescription") or "",
                    "vendor": entry.get("vendorProject") or "",
                    "product_name": entry.get("product") or "",
                    "severity": Vulnerability.Severity.HIGH,
                    "in_kev": True,
                    "kev_data": entry,
                    "modified_at": timezone.now(),
                },
            )
            if not created:
                obj.in_kev = True
                obj.kev_data = entry
                if entry.get("vendorProject") and not obj.vendor:
                    obj.vendor = entry.get("vendorProject") or ""
                if entry.get("product") and not obj.product_name:
                    obj.product_name = entry.get("product") or ""
                obj.save(
                    update_fields=["in_kev", "kev_data", "vendor", "product_name", "updated_at"]
                )
            synced += 1
            if synced % 50 == 0:
                state.mark_progress(synced, len(vulns))
        state.mark_progress(synced, len(vulns))
        state.mark_success(checkpoint={"count": synced})
        return {"ok": True, "synced": synced}
    except Exception as exc:  # noqa: BLE001
        logger.exception("sync_kev failed")
        state.mark_error(str(exc))
        if not Vulnerability.objects.exists():
            seed_demo_cves()
        return {"ok": False, "error": str(exc)}


@shared_task(name="vulndb.apps.vulns.tasks.sync_nvd")
def sync_nvd() -> dict[str, Any]:
    """
    Синхронизация NVD. В конце (успех / demo / exception) при kev_enabled
    синхронно вызывает sync_kev().
    """
    state = SyncState.get_or_create_source(SyncState.Source.NVD)
    settings = _settings()
    result: dict[str, Any] = {"ok": False}
    try:
        state.mark_running(total=NVD_MAX_RECORDS)
        headers = {}
        api_key = settings.nvd_api_key or ""
        import os

        api_key = api_key or os.environ.get("NVD_API_KEY", "")
        if api_key:
            headers["apiKey"] = api_key

        checkpoint = state.checkpoint or {}
        start_index = int(checkpoint.get("startIndex") or 0)
        synced = 0
        total_results = NVD_MAX_RECORDS

        while synced < NVD_MAX_RECORDS:
            params: dict[str, Any] = {
                "resultsPerPage": NVD_PAGE_SIZE,
                "startIndex": start_index,
            }
            # incremental window if we have last success
            if state.last_success_at and checkpoint.get("mode") == "incremental":
                end = timezone.now()
                start = state.last_success_at - timedelta(hours=1)
                params["lastModStartDate"] = start.strftime("%Y-%m-%dT%H:%M:%S.000")
                params["lastModEndDate"] = end.strftime("%Y-%m-%dT%H:%M:%S.000")

            resp = request_with_retries("GET", NVD_URL, timeout=30, headers=headers, params=params)
            if resp.status_code in (403, 404) or not resp.content:
                raise RuntimeError(f"NVD HTTP {resp.status_code}: empty or forbidden")
            resp.raise_for_status()
            payload = resp.json()
            vulns = payload.get("vulnerabilities") or []
            total_results = int(payload.get("totalResults") or len(vulns))
            if not vulns:
                break
            for item in vulns:
                _upsert_cve_from_nvd(item)
                synced += 1
                if synced >= NVD_MAX_RECORDS:
                    break
            state.mark_progress(synced, min(total_results, NVD_MAX_RECORDS))
            start_index += len(vulns)
            if start_index >= total_results:
                break
            # rate-limit pause without API key
            time.sleep(0.7 if api_key else 6.0)

        if synced == 0 and not Vulnerability.objects.exists():
            seeded = seed_demo_cves()
            state.mark_progress(seeded, seeded)
            state.mark_success(checkpoint={"startIndex": 0, "mode": "demo", "demo": True})
            result = {"ok": True, "synced": seeded, "demo": True}
        else:
            state.mark_success(
                checkpoint={"startIndex": start_index, "mode": "incremental", "synced": synced}
            )
            result = {"ok": True, "synced": synced}
    except Exception as exc:  # noqa: BLE001
        logger.exception("sync_nvd failed")
        state.mark_error(str(exc))
        if not Vulnerability.objects.exists():
            seeded = seed_demo_cves()
            result = {"ok": False, "error": str(exc), "demo": seeded}
        else:
            result = {"ok": False, "error": str(exc)}
    finally:
        # RULE: always call sync_kev at the end if kev_enabled
        try:
            if _settings().kev_enabled:
                kev_result = sync_kev()
                result["kev"] = kev_result
        except Exception as kev_exc:  # noqa: BLE001
            logger.exception("sync_kev after nvd failed")
            result["kev_error"] = str(kev_exc)
    return result


def _parse_bdu_row(cells: list) -> dict[str, Any] | None:
    """Parse BDU row by column position (header on row 3)."""
    # Typical BDU columns (0-based):
    # 0 id, 1 title, 2 description, 3 vendor, 4 product, 5 version,
    # 6 vuln_type/status, 7 other_ids, 8 conf/severity?, 9 remediation, ...
    if not cells or len(cells) < 3:
        return None
    bdu_id = str(cells[0] or "").strip()
    if not bdu_id or bdu_id.lower().startswith("идентификатор"):
        return None
    title = str(cells[1] or "").strip()
    description = str(cells[2] or "").strip() if len(cells) > 2 else ""
    vendor = str(cells[3] or "").strip() if len(cells) > 3 else ""
    product = str(cells[4] or "").strip() if len(cells) > 4 else ""
    version = str(cells[5] or "").strip() if len(cells) > 5 else ""
    other_ids = str(cells[7] or "").strip() if len(cells) > 7 else ""
    remediation = str(cells[9] or "").strip() if len(cells) > 9 else ""
    if len(cells) > 12 and not remediation:
        remediation = str(cells[12] or "").strip()
    return {
        "bdu_id": bdu_id,
        "title": title,
        "description": description,
        "vendor": vendor,
        "product_name": product,
        "product_version": version,
        "other_ids": other_ids,
        "remediation": remediation,
        "raw": {f"c{i}": (str(c) if c is not None else "") for i, c in enumerate(cells[:20])},
    }


def _extract_cves(other_ids: str) -> list[str]:
    import re

    return re.findall(r"CVE-\d{4}-\d{4,}", other_ids or "", flags=re.IGNORECASE)


@shared_task(name="vulndb.apps.vulns.tasks.sync_bdu")
def sync_bdu() -> dict[str, Any]:
    state = SyncState.get_or_create_source(SyncState.Source.BDU)
    settings = _settings()
    state.mark_running()
    try:
        import urllib3
        from openpyxl import load_workbook

        if not settings.bdu_verify_ssl:
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

        resp = request_with_retries(
            "GET",
            settings.bdu_xlsx_url,
            timeout=120,
            verify=bool(settings.bdu_verify_ssl),
            stream=True,
        )
        resp.raise_for_status()
        content = resp.content
        # Validate rough size/type
        if len(content) > 80 * 1024 * 1024:
            raise RuntimeError("BDU XLSX слишком большой (>80MB)")
        if len(content) < 100:
            raise RuntimeError("BDU XLSX пустой")

        wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        synced = 0
        total_est = ws.max_row or 0
        state.mark_progress(0, max(total_est - 3, 0))

        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx <= 3:
                continue  # header on row 3
            cells = list(row)
            parsed = _parse_bdu_row(cells)
            if not parsed:
                continue
            cves = _extract_cves(parsed["other_ids"])
            raw = parsed["raw"]
            if cves:
                for cve_id in cves:
                    cve_id = cve_id.upper()
                    obj, created = Vulnerability.objects.get_or_create(
                        vuln_id=cve_id,
                        defaults={
                            "record_type": Vulnerability.RecordType.CVE,
                            "title": parsed["title"] or cve_id,
                            "description_bdu": parsed["description"],
                            "vendor": parsed["vendor"],
                            "product_name": parsed["product_name"],
                            "product_version": parsed["product_version"],
                            "remediation": parsed["remediation"],
                            "bdu_id": parsed["bdu_id"],
                            "has_bdu": True,
                            "bdu_raw": raw,
                            "modified_at": timezone.now(),
                        },
                    )
                    if not created:
                        obj.description_bdu = parsed["description"] or obj.description_bdu
                        obj.vendor = parsed["vendor"] or obj.vendor
                        obj.product_name = parsed["product_name"] or obj.product_name
                        obj.product_version = parsed["product_version"] or obj.product_version
                        obj.remediation = parsed["remediation"] or obj.remediation
                        obj.bdu_id = parsed["bdu_id"]
                        obj.has_bdu = True
                        obj.bdu_raw = raw
                        obj.save()
            else:
                vuln_id = f"BDU:{parsed['bdu_id']}"
                Vulnerability.objects.update_or_create(
                    vuln_id=vuln_id,
                    defaults={
                        "record_type": Vulnerability.RecordType.BDU,
                        "title": parsed["title"] or vuln_id,
                        "description_bdu": parsed["description"],
                        "vendor": parsed["vendor"],
                        "product_name": parsed["product_name"],
                        "product_version": parsed["product_version"],
                        "remediation": parsed["remediation"],
                        "bdu_id": parsed["bdu_id"],
                        "has_bdu": True,
                        "bdu_raw": raw,
                        "modified_at": timezone.now(),
                    },
                )
            synced += 1
            if synced % 100 == 0:
                state.mark_progress(synced, max(total_est - 3, synced))

        wb.close()
        state.mark_progress(synced, synced)
        state.mark_success(checkpoint={"synced": synced})
        return {"ok": True, "synced": synced}
    except Exception as exc:  # noqa: BLE001
        logger.exception("sync_bdu failed")
        state.mark_error(str(exc))
        if not Vulnerability.objects.exists():
            seed_demo_cves()
        return {"ok": False, "error": str(exc)}


@shared_task(name="vulndb.apps.vulns.tasks.tick_sync_schedules")
def tick_sync_schedules() -> dict[str, Any]:
    """Every 60s: launch due syncs according to intervals and last_success_at."""
    settings = _settings()
    now = timezone.now()
    launched: list[str] = []

    def due(enabled: bool, interval_min: int, source: str) -> bool:
        if not enabled:
            return False
        st = SyncState.get_or_create_source(source)
        if st.status == SyncState.Status.RUNNING:
            return False
        if not st.last_success_at:
            return True
        return (now - st.last_success_at) >= timedelta(minutes=interval_min)

    if due(settings.nvd_enabled, settings.nvd_sync_interval_minutes, SyncState.Source.NVD):
        sync_nvd.delay()
        launched.append("nvd")
    elif due(
        settings.kev_enabled and not settings.nvd_enabled,
        settings.kev_sync_interval_minutes,
        SyncState.Source.KEV,
    ):
        # KEV separately only when NVD disabled
        sync_kev.delay()
        launched.append("kev")

    if due(settings.bdu_enabled, settings.bdu_sync_interval_minutes, SyncState.Source.BDU):
        sync_bdu.delay()
        launched.append("bdu")

    return {"launched": launched}
